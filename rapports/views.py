from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import F, Sum
from ventes.models import Vente
from produits.models import Produit
from clients.models import Client
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from utilisateurs.decorators import gerant_required


@login_required
@gerant_required
def rapports_index(request):
    return render(request, 'rapports/index.html')


@login_required
@gerant_required
def rapport_ventes(request):
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    ventes = Vente.objects.all()
    if date_debut:
        ventes = ventes.filter(date_vente__gte=date_debut)
    if date_fin:
        ventes = ventes.filter(date_vente__lte=date_fin)
    
    total_ventes = ventes.aggregate(Sum('montant_total'))['montant_total__sum'] or 0
    total_paye = ventes.aggregate(Sum('montant_paye'))['montant_paye__sum'] or 0
    nb_ventes = ventes.count()
    
    context = {
        'ventes': ventes,
        'total_ventes': total_ventes,
        'total_paye': total_paye,
        'nb_ventes': nb_ventes,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    return render(request, 'rapports/ventes.html', context)


@login_required
@gerant_required
def rapport_stock(request):
    produits = Produit.objects.all()
    
    total_stock_value = sum(p.stock_actuel * p.prix_achat for p in produits)
    produits_alerte = produits.filter(stock_actuel__lte=F('seuil_alerte'))
    
    context = {
        'produits': produits,
        'total_stock_value': total_stock_value,
        'produits_alerte': produits_alerte,
    }
    return render(request, 'rapports/stock.html', context)


@login_required
@gerant_required
def export_excel_ventes(request):
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    ventes = Vente.objects.all()
    if date_debut:
        ventes = ventes.filter(date_vente__gte=date_debut)
    if date_fin:
        ventes = ventes.filter(date_vente__lte=date_fin)
    
    # Créer le classeur
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventes"
    
    # En-têtes
    headers = ['Numéro', 'Client', 'Date', 'Montant Total', 'Montant Payé', 'Solde', 'Statut']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    
    # Données
    for row, vente in enumerate(ventes, 2):
        ws.cell(row=row, column=1, value=vente.numero)
        ws.cell(row=row, column=2, value=vente.client.nom if vente.client else "Anonyme")
        ws.cell(row=row, column=3, value=vente.date_vente)
        ws.cell(row=row, column=4, value=vente.montant_total)
        ws.cell(row=row, column=5, value=vente.montant_paye)
        ws.cell(row=row, column=6, value=vente.solde_restant)
        ws.cell(row=row, column=7, value=vente.get_statut_display())
    
    # Ajuster les largeurs
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    # Retourner le fichier
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="rapport_ventes.xlsx"'
    wb.save(response)
    return response
